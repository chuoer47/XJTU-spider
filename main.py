# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from config import *  # 配置文件中填写相关参数
from util import *


def spider(flag: int,
           name: str,
           use_suffix=False,
           begin_url='http://dean.xjtu.edu.cn/jxzy/jpkc',
           ) -> None:
    while flag:
        # 构造链接
        if use_suffix:
            start_url = begin_url + '/' + str(flag) + '.htm'
        else:
            start_url = begin_url + '.htm'

        # 初始化 DataFrame
        df_all = pd.DataFrame(columns=['标题', '发布日期', '浏览次数', '正文内容'])

        # 访问起始页面并提取链接
        links_to_visit = extract_links(start_url)

        # 顺序访问链接并提取数据
        for link in links_to_visit:
            html = get_full_html(link)
            if html:
                title, release_date, view_count, body_text = extract_info(html, link)
                df_all = pd.concat([df_all, pd.DataFrame(
                    {'标题': [title], '发布日期': [release_date], '浏览次数': [view_count],
                     '正文内容': [body_text]})], ignore_index=True)

        # 读取已存在的 Excel 文件或新建一个
        try:
            wb = load_workbook(datafile + name + '.xlsx')
            ws = wb.active
            row_count = ws.max_row
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            row_count = 1

        # 将 DataFrame 中的数据逐行写入 Excel 文件中
        for index, row in df_all.iterrows():
            ws.cell(row=row_count + index, column=1, value=row['标题'])
            ws.cell(row=row_count + index, column=2, value=row['发布日期'])
            ws.cell(row=row_count + index, column=3, value=row['浏览次数'])
            ws.cell(row=row_count + index, column=4, value=row['正文内容'])

        # 保存 Excel 文件
        wb.save(datafile + name + '.xlsx')
        print("已保存" + str(flag) + "页信息")

        # 更新 flag1 的值
        flag -= 1
        use_suffix = True  # 开始使用后缀

    print("所有信息已保存到" + name + ".xlsx 文件中。")


if __name__ == "__main__":
    # 下面是爬取示例程序
    flag = 30  # 初始化 flag 为爬取页面的总页数-1
    name = "精品课程"  # 修改项
    use_suffix = False  # 是否在 start_url 中添加后缀的标志
    begin_url = 'http://dean.xjtu.edu.cn/jxzy/jpkc'  # 'http://dean.xjtu.edu.cn/jxzy/jpkc.hml'的前缀
    spider(flag=flag,
           name=name,
           use_suffix=use_suffix,
           begin_url=begin_url)
    print("爬取完毕！")
